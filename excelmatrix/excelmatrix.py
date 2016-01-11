#!/usr/bin/env python

__author__ = "Stefan Lieberth"
__copyright__ = "Copyright 2015"
__credits__ = [""]
__license__ = "MIT/Expat"
__version__ = "1.5"
__maintainer__ = "Stefan Lieberth"
__email__ = "stefan@lieberth.net"

import os.path
from pprint import pprint
from openpyxl import load_workbook, Workbook
from openpyxl.compat import range
from openpyxl.utils import (
    coordinate_from_string,
    COORD_RE,
    ABSOLUTE_RE,
    column_index_from_string,
    get_column_letter,
    range_boundaries,
    rows_from_range,
    coordinate_to_tuple,
)
from openpyxl.styles import Style, PatternFill, Border, Side, Alignment, Protection, Font,Color


def writeMatrix (*args,**kwargs):
    return excelmatrix._writeMatrix(*args,**kwargs)

def readMatrix (*args,**kwargs):
    return excelmatrix._readMatrix(*args,**kwargs)

class excelmatrix:

    
    def __init__(self):
        pass
    
    @classmethod
    def _writeMatrix (self,xlsxFilename,xlsxSheetname,contentMatrix,**kwargs):


        def checkFileAndSheet():
            if os.path.isfile(xlsxFilename):
                outWorkbook = load_workbook(xlsxFilename)
                if xlsxSheetname in outWorkbook.get_sheet_names():
                    outSheet = outWorkbook[xlsxSheetname]
                else:
                    outSheet = outWorkbook.create_sheet()
                    outSheet.title = xlsxSheetname
            else:
                outWorkbook = Workbook()
                outWorkbook.remove_sheet(outWorkbook.worksheets[0])
                outSheet = outWorkbook.create_sheet()
                outSheet.title = xlsxSheetname
            return outWorkbook, outSheet

        def writeLine(lineRow,columnOffset,contentList):
            for myColumn in range (1,len(contentList) +1 ):
                myCell = self.outSheet['%s%s'%(get_column_letter(myColumn + columnOffset), lineRow)]
                #print contentList[myColumn - 1 ]   ###FIXME###
                myCell.value = contentList[myColumn - 1 ]
                #myCell.style = cellStyle

        self.outWorkbook, self.outSheet = checkFileAndSheet()

        if 'start' in kwargs.keys():
            startCell = kwargs["start"] 
            startColumn, startRow = coordinate_from_string (startCell)
            rowOffset = startRow - 1
            columnOffset = column_index_from_string(startColumn) - 1
        else: 
            rowOffset = 0
            columnOffset = 0

        for myRowIndex in range (1 ,len(contentMatrix) +1 ):
                writeLine (rowOffset + myRowIndex,columnOffset,contentMatrix[myRowIndex - 1])
        self.outWorkbook.save(filename = xlsxFilename)
        return True


    @classmethod
    def _readMatrix (self,xlsxFilename,xlsxSheetname,**kwargs):
        workBook = load_workbook(xlsxFilename)
        inSheet = workBook[xlsxSheetname]
        returnMatrix = []
        if 'range' in kwargs.keys():
            rangeString = kwargs["range"] 
            startRange, endRange = rangeString.split (":")
            startColumn, startRow = coordinate_from_string (startRange)
            endColumn, endRow = coordinate_from_string (endRange)
            for row in range (startRow,endRow + 1):
                lineArray = []
                for column in range (column_index_from_string(startColumn),column_index_from_string(endColumn) + 1):
                    #print inSheet['%s%s'%(get_column_letter(column), row)].value
                    lineArray.append(inSheet['%s%s'%(get_column_letter(column), row)].value)
                    column += 1
                returnMatrix.append(lineArray)
                column = 1
                row += 1
            return returnMatrix
        else:
            if 'start' in kwargs.keys():
                startCell = kwargs["start"] 
                startColumnLetter, startRow = coordinate_from_string (startCell)
                startColumn = column_index_from_string(startColumnLetter)
            else:
                startColumn, startRow = 1,1  
            if 'cellsPerRow' in kwargs.keys():
                numberOfCellsPerRow = kwargs["cellsPerRow"]      
            row = startRow
            column = startColumn
            while inSheet['%s%s'%(get_column_letter(column), row)].value:
                lineArray = []
                if numberOfCellsPerRow: 
                    while column <= startColumn + numberOfCellsPerRow - 1:
                        #print inSheet['%s%s'%(get_column_letter(column), row)].value
                        lineArray.append(inSheet['%s%s'%(get_column_letter(column), row)].value)
                        column += 1
                else:
                    while inSheet['%s%s'%(get_column_letter(column), row)].value:
                        #print inSheet['%s%s'%(get_column_letter(column), row)].value
                        lineArray.append(inSheet['%s%s'%(get_column_letter(column), row)].value)
                        column += 1
                returnMatrix.append(lineArray)
                column = startColumn
                row += 1
            return returnMatrix






